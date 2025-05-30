import base64
import json
import logging
import math
import re
import uuid
from io import BytesIO

import pandas as pd
from django.conf import settings
from django.db import models
from django.urls import reverse
from django.utils import timezone
from django_extensions.db.models import TimeStampedModel
from jinja2 import DebugUndefined, Environment
from openpyxl import Workbook
from openpyxl.styles import Alignment

from actionable.mixins import ActionableObjectMixin, action_with_permission

from .extract import extract_text
from .langfuse_api import cost_for_session_usd
from .llm_calling import chatter

logger = logging.getLogger(__name__)


class Tool(models.Model):
    name = models.CharField(max_length=100)
    prompt = models.TextField(
        help_text="Use {{curly}} braces for the source placeholder, and square brackets for LLM completions, e.g. [[response]]",
        default="""{{source}}\n\nQuestion about the source\n\n[[response]]""",
    )
    model = models.ForeignKey(
        "mindframe.LLM", on_delete=models.CASCADE, help_text="The LLM to use for this Tool"
    )
    include_source = models.BooleanField(
        default=False, help_text="Include the source text in the results JSON (can be quite large)."
    )

    def __str__(self):
        return self.name

    def get_input_fields(self):
        # finds all {placeholder_name} in prompt
        return re.findall(r"\{\{([^}]+)\}\}", self.prompt)

    def get_absolute_url(self):
        return reverse("tool_input", kwargs={"pk": self.pk})


class ToolKey(models.Model):
    """
    This model links a Tool with a secure UUID key
    so that clients can post data to the Ninja API with that key.
    """

    tool = models.ForeignKey(Tool, on_delete=models.CASCADE, related_name="tool_keys")
    tool_key = models.UUIDField(default=uuid.uuid4, unique=True)

    def __str__(self):
        return f"{self.tool.name} - {self.tool_key}"


class JobGroup(ActionableObjectMixin, TimeStampedModel):
    uuid = models.UUIDField(default=uuid.uuid4, unique=True)
    tool = models.ForeignKey(Tool, on_delete=models.CASCADE)
    cancelled = models.BooleanField(default=False)
    owner = models.ForeignKey(
        settings.AUTH_USER_MODEL, on_delete=models.CASCADE, related_name="job_groups"
    )

    def cost(self):
        tc = cost_for_session_usd(self.uuid)
        return tc and math.ceil(tc * 1000) / 1000 or None

    def complete(self):
        return all(self.jobs.all().values_list("completed", flat=True))

    def in_progress(self):
        return (not self.cancelled) and (not self.complete)

    def status(self):
        if self.cancelled:
            return "Cancelled"
        elif self.complete:
            return "Complete"
        else:
            return "In progress"

    def get_absolute_url(self):
        return reverse("jobgroup-detail", kwargs={"pk": self.pk})

    def results(self):
        return list(self.jobs.all().values_list("result", flat=True))

    def csv(self):
        return pd.DataFrame(self.results()).to_csv()

    def xlsx(self):
        data = list(self.jobs.all().values_list("result", flat=True))
        wb = Workbook()
        ws = wb.active

        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)

        # Write rows
        for row in data:
            ws.append([row[h] for h in headers])

        # Enable text wrap for all cells
        wrap_alignment = Alignment(wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = wrap_alignment

        # Optionally auto-adjust column width
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[col[0].column_letter].width = adjusted_width

        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def json(self):
        return json.dumps(list(self.jobs.all().values_list("result", flat=True)), indent=2)

    @action_with_permission("llmtools.cancel_jobgroup")
    def cancel_jobs(self):
        self.jobs.update(cancelled=True)
        self.cancelled = True
        self.save()
        return True

    def run(self):
        if self.cancelled:
            raise Exception("Job group cancelled")

        logger.info(f"Processing JobGroup: {self.id}")
        for job in self.jobs.all():
            job.run()

        if self.jobs.filter(completed__isnull=True).count() == 0:
            self.complete = True
            self.save()
            for i in self.jobs.all():
                i.source_file.delete()

    class Meta:
        ordering = ["-created"]


class Job(TimeStampedModel):
    group = models.ForeignKey(JobGroup, on_delete=models.CASCADE, related_name="jobs", null=True)
    tool = models.ForeignKey(Tool, on_delete=models.CASCADE, null=True)
    cancelled = models.BooleanField(default=False)

    def ready_to_run(self):
        if self.group and self.group.cancelled:
            return False

        return not self.cancelled and not self.completed

    def get_tool(self):
        return self.tool or self.group.tool

    context = models.JSONField(null=True)

    source_file = models.FileField(upload_to="source_files", max_length=1024 * 4)

    result = models.JSONField(null=True)
    completed = models.DateTimeField(null=True)

    def __str__(self):
        return f"{self.source_file or str(self.context)[:50]}..."

    def text(self):
        return (
            self.source_file
            and extract_text(self.source_file.path)
            or self.context.get("source", None)
            or str(self.context)
        )

    def filename(self):
        return self.source_file and self.source_file.name or ""

    def filepath(self):
        return self.source_file and self.source_file.path or None

    def run(self, force=False):

        logger.info(f"Running job: {self.id}")

        result = None
        try:
            if self.cancelled:
                raise Exception("Job cancelled")

            if self.completed and not force:
                raise Exception("Job already completed")

            tool = self.get_tool()
            # Create a Template object, use jinja2 for rendering becase it preserves
            # unmatched {{var}} placeholders
            env = Environment(undefined=DebugUndefined)
            template = env.from_string(tool.prompt)

            filled_prompt = template.render(self.context)
            logger.warning(f"Filled prompt: {filled_prompt}")
            result = chatter(filled_prompt, tool.model)
            if len(result.keys()) > 1:
                del result["RESPONSE_"]

            if tool.include_source:
                result["source"] = self.text()

            result["file_name"] = self.filename()

            self.result = result
            self.completed = timezone.now()
            self.save()

        except Exception as e:
            logger.error(f"Error running job: {self.id}: {e}")

        return result
